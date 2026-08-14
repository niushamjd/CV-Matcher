import time
import re
import os
import json
import numpy as np
import pandas as pd
import torch 
from sentence_transformers import SentenceTransformer

class EmbeddingMatcher:
    def __init__(self, model_name: str = "all-MiniLM-L6-v2"):
        # SentenceTransformers handles local dense embeddings quickly with 0 API cost
        torch.set_num_threads(1)
        self.model = SentenceTransformer(model_name, device="cpu")

    def _cosine_similarity(self, vec1: np.ndarray, vec2: np.ndarray) -> float:
        norm1, norm2 = np.linalg.norm(vec1), np.linalg.norm(vec2)
        if norm1 == 0 or norm2 == 0:
            return 0.0
        return float(np.dot(vec1, vec2) / (norm1 * norm2))

    def _extract_sections(self, text: str) -> dict[str, str]:
        """Simple regex chunking to separate CV into key sections."""
        sections = {"skills": "", "experience": "", "education": "", "full": text}
        
        # Lightweight regex headers
        skills_match = re.search(r"(?i)(skills|technical skills|competencies)([\s\S]*?)(?=experience|education|$)", text)
        exp_match = re.search(r"(?i)(experience|work history|employment)([\s\S]*?)(?=skills|education|$)", text)
        edu_match = re.search(r"(?i)(education|academic background)([\s\S]*?)(?=skills|experience|$)", text)

        if skills_match: sections["skills"] = skills_match.group(2).strip()
        if exp_match: sections["experience"] = exp_match.group(2).strip()
        if edu_match: sections["education"] = edu_match.group(2).strip()
        
        return sections

    def match(self, cv_text: str, jd_text: str, section_aware: bool = True) -> dict:
        start_time = time.time()
        
        if not section_aware:
            # Full Document Similarity
            embeddings = self.model.encode([cv_text, jd_text])
            sim = self._cosine_similarity(embeddings[0], embeddings[1])
            sub_scores = {"full_text_sim": round(sim, 4)}
        else:
            # Section-Aware Similarity
            cv_sec = self._extract_sections(cv_text)
            jd_sec = self._extract_sections(jd_text)
            
            # Embed all parts in a single batch for speed
            texts_to_embed = [
                cv_text, jd_text,
                cv_sec["skills"] or cv_text, jd_sec["skills"] or jd_text,
                cv_sec["experience"] or cv_text, jd_sec["experience"] or jd_text
            ]
            vecs = self.model.encode(texts_to_embed)
            
            full_sim = self._cosine_similarity(vecs[0], vecs[1])
            skills_sim = self._cosine_similarity(vecs[2], vecs[3])
            exp_sim = self._cosine_similarity(vecs[4], vecs[5])
            
            # Weighted combine (40% skills, 40% experience, 20% overall)
            sim = (0.4 * skills_sim) + (0.4 * exp_sim) + (0.2 * full_sim)
            
            sub_scores = {
                "overall_sim": round(full_sim, 4),
                "skills_sim": round(skills_sim, 4),
                "experience_sim": round(exp_sim, 4)
            }

        match_score = round(float(sim) * 100, 2)
        latency = (time.time() - start_time) * 1000

        
        return {
            "method_name": "2a_embedding_similarity",
            "match_score": match_score,
            "sub_scores": sub_scores,
            "explanation": f"Dense vector similarity score of {match_score}%.",
            "latency_ms": round(latency, 2),
            "total_tokens": 0
        }


if __name__ == "__main__":
    from pathlib import Path
    import openpyxl

    _REPO_ROOT = Path(__file__).resolve().parent.parent.parent
    EXCEL_PATH = _REPO_ROOT / "data" / "cv_matcher_candidate_pool.xlsx"
    RESULTS_PATH = _REPO_ROOT / "results" / "embedding_results.json"

    # Load gold-standard pairs (mirrors run_on_gold_standard.py)
    wb = openpyxl.load_workbook(str(EXCEL_PATH), data_only=True)
    ws = wb["Candidate Pool"]
    header = [c.value for c in ws[1]]
    pairs = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = dict(zip(header, row))
        if not (record.get("resume_text") and record.get("jd_text")):
            continue
        if record.get("FINAL_SELECTION (y/n)") != "y":
            continue
        pairs.append(record)
    print(f"Loaded {len(pairs)} gold-standard pairs.")

    # Resume: skip pairs already done
    results = {}
    if RESULTS_PATH.exists():
        with open(RESULTS_PATH, encoding="utf-8") as f:
            results = json.load(f)
        print(f"Resuming — {len(results)} pairs already done.\n")

    matcher = EmbeddingMatcher()

    for record in pairs:
        pair_id = str(record.get("pair_id", ""))
        if pair_id in results:
            print(f"  {pair_id} — already done, skipping")
            continue

        cv_text = str(record["resume_text"]).strip()
        jd_text = str(record["jd_text"]).strip()

        res = matcher.match(cv_text, jd_text, section_aware=True)
        results[pair_id] = res
        print(f"  {pair_id}: match_score={res['match_score']}")

        RESULTS_PATH.parent.mkdir(parents=True, exist_ok=True)
        with open(RESULTS_PATH, "w", encoding="utf-8") as f:
            json.dump(results, f, indent=2)

    print(f"\nSaved {len(results)} results to {RESULTS_PATH}")
