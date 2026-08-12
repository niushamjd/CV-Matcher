"""
Data loader for CV data in two formats:

  1. CSV (Kaggle "Resume Dataset"):  columns ID, Resume_str, Resume_html, Category
     → use load_resumes_csv()

  2. PDF folder layout:  data/data/<CATEGORY>/<id>.pdf
     → use load_resumes_pdf()

Both return a list[ResumeRecord] with the same shape so the rest of the
pipeline (extraction.py etc.) is identical regardless of source.
"""

import glob
import os
from dataclasses import dataclass
from pathlib import Path
from typing import Optional

import pandas as pd

# Repo root is one level above this file (src/)
_REPO_ROOT = Path(__file__).resolve().parent.parent.parent  # methods/structured_extraction/ -> methods/ -> repo root

# Default paths — override by passing explicit arguments if your layout differs
DEFAULT_CSV_PATH = _REPO_ROOT / "Resume" / "Resume.csv"
DEFAULT_PDF_ROOT = _REPO_ROOT / "data" / "data"
DEFAULT_GOLD_EXCEL = _REPO_ROOT / "annotation" / "cv_matcher_candidate_pool.xlsx"


@dataclass
class ResumeRecord:
    id: str
    category: str
    text: str


def _clean_resume_text(raw: str) -> str:
    if not isinstance(raw, str):
        return ""
    return " ".join(raw.replace("\r", " ").split())


# ---------------------------------------------------------------------------
# 1. CSV loader (Kaggle Resume Dataset)
# ---------------------------------------------------------------------------

def load_resumes_csv(
    csv_path: Optional[str] = None,
    category: Optional[str] = None,
    n: Optional[int] = None,
) -> list[ResumeRecord]:
    """
    Load resumes from the Kaggle CSV (ID, Resume_str, Resume_html, Category).

    csv_path: defaults to Resume/Resume.csv relative to the repo root.
    category: filter to one category, e.g. "BANKING" (case-insensitive).
    n: cap at first n rows after filtering.
    """
    path = str(csv_path or DEFAULT_CSV_PATH)
    df = pd.read_csv(path)

    missing = {"ID", "Resume_str", "Category"} - set(df.columns)
    if missing:
        raise ValueError(f"CSV at {path} is missing columns: {missing}")

    if category:
        df = df[df["Category"].str.upper() == category.upper()]
    if n:
        df = df.head(n)

    return [
        ResumeRecord(
            id=str(row["ID"]),
            category=str(row["Category"]),
            text=_clean_resume_text(row["Resume_str"]),
        )
        for _, row in df.iterrows()
    ]


# ---------------------------------------------------------------------------
# 2. PDF loader (data/data/<CATEGORY>/<id>.pdf)
# ---------------------------------------------------------------------------

def _pdf_to_text(pdf_path: str) -> str:
    """Extract plain text from a PDF using pdfplumber."""
    try:
        import pdfplumber
    except ImportError:
        raise ImportError("pdfplumber is required for PDF loading. Run: pip install pdfplumber")

    text_parts = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text()
            if page_text:
                text_parts.append(page_text)
    return _clean_resume_text("\n".join(text_parts))


def load_resumes_pdf(
    root_dir: Optional[str] = None,
    category: Optional[str] = None,
    n: Optional[int] = None,
) -> list[ResumeRecord]:
    """
    Load resumes from PDF files under data/data/<CATEGORY>/<id>.pdf.

    root_dir: defaults to data/data/ relative to the repo root.
    category: load only that subfolder, e.g. "BANKING".
    n: cap at first n PDFs (per category if no category filter, total if filtered).
    """
    root = Path(root_dir or DEFAULT_PDF_ROOT)

    if category:
        search_dirs = [root / category.upper()]
    else:
        search_dirs = sorted(p for p in root.iterdir() if p.is_dir())

    records: list[ResumeRecord] = []
    for cat_dir in search_dirs:
        cat_name = cat_dir.name
        pdfs = sorted(cat_dir.glob("*.pdf"))
        if n and not category:
            pdfs = pdfs[:n]

        for pdf_path in pdfs:
            text = _pdf_to_text(str(pdf_path))
            if text:
                records.append(
                    ResumeRecord(
                        id=pdf_path.stem,
                        category=cat_name,
                        text=text,
                    )
                )

    if n and category:
        records = records[:n]

    return records


# ---------------------------------------------------------------------------
# 3. Gold-standard Excel loader (annotation/cv_matcher_candidate_pool.xlsx)
# ---------------------------------------------------------------------------

@dataclass
class GoldPair:
    pair_id: str
    resume_text: str
    jd_text: str


def load_pairs_from_excel(
    excel_path: Optional[str] = None,
    sheet_name: str = "Candidate Pool",
    final_only: bool = True,
) -> list[GoldPair]:
    """
    Load the 30 gold-standard CV–JD pairs from Buse's annotation Excel.
    Mirrors Niyousha's load_pairs_from_excel() so both methods run on identical input.

    excel_path: defaults to annotation/cv_matcher_candidate_pool.xlsx
    final_only: if True (default), only rows where FINAL_SELECTION (y/n) == "y"
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl is required. Run: pip install openpyxl")

    path = str(excel_path or DEFAULT_GOLD_EXCEL)
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet_name]
    header = [c.value for c in ws[1]]

    pairs = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = dict(zip(header, row))
        if not (record.get("resume_text") and record.get("jd_text")):
            continue
        if final_only and record.get("FINAL_SELECTION (y/n)") != "y":
            continue
        pairs.append(GoldPair(
            pair_id=str(record.get("pair_id", "")),
            resume_text=_clean_resume_text(str(record["resume_text"])),
            jd_text=str(record["jd_text"]).strip(),
        ))
    return pairs


if __name__ == "__main__":
    # --- Test CSV loader ---
    print("=== CSV loader ===")
    csv_records = load_resumes_csv(category="BANKING", n=2)
    for r in csv_records:
        print(f"{r.id} ({r.category}): {r.text[:150]}...\n")

    # --- Test PDF loader ---
    print("=== PDF loader ===")
    pdf_records = load_resumes_pdf(category="BANKING", n=2)
    for r in pdf_records:
        print(f"{r.id} ({r.category}): {r.text[:150]}...\n")
