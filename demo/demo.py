"""
Live demo for the project video: runs all three CV-Matcher methods
(2a Dense Embedding Similarity, 2b Structured Extraction + Rules,
2c LLM-as-Judge) on a single gold-standard pair, narrates each step as
it happens, and ends with a side-by-side summary table against the
human annotation.

Pair: PAIR_13 (Senior Human Resources Specialist) -- a clean strong-fit
case where all three methods produce presentable, non-degenerate output.

Setup (see README.md "API Keys"):
    GEMINI_API_KEY   required (method 2c, and method 2b if LLM_PROVIDER=gemini)
    GROQ_API_KEY     required for method 2b's default provider (LLM_PROVIDER=groq)

Run from the repo root:
    python demo/demo.py
"""

import sys
import time
from pathlib import Path

from dotenv import load_dotenv
load_dotenv()

REPO_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO_ROOT / "methods" / "structured_extraction"))
sys.path.insert(0, str(REPO_ROOT / "methods" / "embedding_similarity"))
sys.path.insert(0, str(REPO_ROOT / "methods" / "llm_judge"))

import openpyxl

PAIR_ID = "PAIR_13"


def load_pair(pair_id: str) -> tuple[str, str]:
    wb = openpyxl.load_workbook(REPO_ROOT / "data" / "cv_matcher_candidate_pool.xlsx", data_only=True)
    ws = wb["Candidate Pool"]
    header = [c.value for c in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = dict(zip(header, row))
        if str(record.get("pair_id")) == pair_id:
            return str(record["resume_text"]).strip(), str(record["jd_text"]).strip()
    raise ValueError(f"{pair_id} not found in {REPO_ROOT / 'data' / 'cv_matcher_candidate_pool.xlsx'}")


def load_human_label(pair_id: str) -> dict:
    import pandas as pd
    df = pd.read_excel(REPO_ROOT / "data" / "annotation" / "annotated_pairs.xlsx", sheet_name="Sheet1")
    row = df[df["pair_id"].astype(str) == pair_id].iloc[0]
    return {
        "overall_fit_1to5": row["overall_fit_1to5"],
        "skills_1to5": row["subscore_skills_1to5"],
        "experience_1to5": row["subscore_experience_1to5"],
        "education_1to5": row["subscore_education_1to5"],
        "shortlist_decision": row["shortlist_decision"],
        "rationale": row["rationale"],
    }


def section(title: str):
    print("\n" + "=" * 72)
    print(title)
    print("=" * 72)


def step(msg: str):
    """One-line narration printed right before a slow/API call, so the
    terminal doesn't just sit there silently while a request is in flight."""
    print(f"  -> {msg}")


def run_2a(cv_text: str, jd_text: str) -> dict:
    from embedding_matcher import EmbeddingMatcher

    step("Loading all-MiniLM-L6-v2 locally (SentenceTransformers, CPU, no API call)...")
    matcher = EmbeddingMatcher()

    step("Regex-extracting Skills/Experience sections from CV and JD...")
    step("Encoding sections into embeddings and computing weighted cosine similarity "
         "(40% skills, 40% experience, 20% overall)...")
    result = matcher.match(cv_text, jd_text, section_aware=True)

    print(f"match_score : {result['match_score']}")
    print(f"sub_scores  : {result['sub_scores']}")
    print(f"latency_ms  : {result['latency_ms']}")

    return {
        "label": "2a Dense Embedding",
        "match_score": result["match_score"],
        "latency_s": round(result["latency_ms"] / 1000, 2),
    }


def run_2b(cv_text: str, jd_text: str) -> dict:
    from extraction import extract_profile
    from rule_scorer import score_match
    from gap_explainer import explain_llm, explain_template
    from llm_client import writing_quality_signal

    start = time.time()

    step("Calling LLM (extraction) to parse the CV into a structured profile "
         "(skills, experience, education)...")
    cv_profile = extract_profile(cv_text, source_type="cv")

    step("Calling LLM (extraction) to parse the JD into the same structured schema...")
    jd_profile = extract_profile(jd_text, source_type="jd")

    step("Scoring skills coverage, experience gap, and education match "
         "deterministically -- no LLM call for this step...")
    result = score_match(cv_profile, jd_profile)

    step("Calling LLM for a writing-quality signal (clarity/professionalism only, "
         "kept separate from match_score)...")
    wq = writing_quality_signal(cv_text)
    result.sub_scores.writing_quality_score = wq["writing_quality_score"]

    step("Calling LLM (gap explainer) to turn the structured diff into a "
         "natural-language rationale...")
    try:
        result.explanation = explain_llm(result.structured_diff, wq["writing_quality_score"])
    except Exception:
        step("Gap-explainer LLM call failed -- falling back to the deterministic template...")
        result.explanation = explain_template(result.structured_diff)

    latency = round(time.time() - start, 2)
    print(f"match_score : {result.match_score}")
    print(f"sub_scores  : {result.sub_scores.model_dump()}")
    print(f"explanation : {result.explanation}")
    print(f"latency_s   : {latency}")

    return {
        "label": "2b Structured Extraction",
        "match_score": result.match_score,
        "latency_s": latency,
    }


def run_2c(cv_text: str, jd_text: str) -> dict:
    from llm_judge_starter import llm_as_judge

    step("Calling Gemini with the raw CV + JD in a single prompt, grounded in a "
         "two-tier must-have/nice-to-have recruiting rubric...")
    result = llm_as_judge(cv_text, jd_text)

    print(f"match_score : {result['match_score']}")
    print(f"sub_scores  : {result['sub_scores']}")
    print(f"explanation : {result['explanation']}")
    print(f"latency_s   : {result['latency_seconds']}")

    return {
        "label": "2c LLM-as-Judge",
        "match_score": result["match_score"],
        "latency_s": result["latency_seconds"],
    }


def print_summary(human: dict, results: list[dict]):
    human_score = human["overall_fit_1to5"] * 20  # rescale 1-5 -> 0-100

    section("Summary")
    print(f"{'Method':<28}{'Match Score':>13}{'Latency (s)':>14}{'Delta vs Human':>18}")
    print("-" * 73)
    print(f"{'Human gold label':<28}{human_score:>13.1f}{'--':>14}{'--':>18}")
    for r in results:
        if r is None:
            continue
        delta = r["match_score"] - human_score
        print(f"{r['label']:<28}{r['match_score']:>13.2f}{r['latency_s']:>14.2f}{delta:>+18.1f}")
    skipped = [r for r in results if r is None]
    if skipped:
        print(f"\n({len(skipped)} method(s) skipped -- see [skipped] messages above)")
    print()


def main():
    cv_text, jd_text = load_pair(PAIR_ID)
    human = load_human_label(PAIR_ID)

    section(f"CV Matcher demo -- {PAIR_ID}")
    print(f"Human label : overall {human['overall_fit_1to5']}/5, "
          f"skills {human['skills_1to5']}/5, experience {human['experience_1to5']}/5, "
          f"education {human['education_1to5']}/5, decision: {human['shortlist_decision']}")
    print(f"Rationale   : {human['rationale']}")

    results = []

    section("Method 2a -- Dense Embedding Similarity")
    try:
        results.append(run_2a(cv_text, jd_text))
    except Exception as e:
        print(f"[skipped] {e}")
        results.append(None)

    section("Method 2b -- Structured Extraction + Rules")
    try:
        results.append(run_2b(cv_text, jd_text))
    except Exception as e:
        print(f"[skipped] {e}")
        results.append(None)

    section("Method 2c -- LLM-as-Judge")
    try:
        results.append(run_2c(cv_text, jd_text))
    except Exception as e:
        print(f"[skipped] {e}")
        results.append(None)

    print_summary(human, results)


if __name__ == "__main__":
    main()
